import os
import io
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, text

# Import openpyxl components for Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.secret_key = "coolaire_binibini_2026_secret_key"

# --- DATABASE CONFIGURATION ---
app.config[
    'SQLALCHEMY_DATABASE_URI'] = 'postgresql+psycopg2://postgres.tqjvwfikswvppeyopvdg:3srdUc8IFDiUkbJu@aws-1-ap-northeast-1.pooler.supabase.com:6543/postgres'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# --- CATEGORY SCHEMAS ---
CATEGORIES = {
    1: {'name': 'Dept. Representation', 'max': 25, 'field': 'dept_rep'},
    2: {'name': 'Intro & Communication', 'max': 25, 'field': 'comm_skills'},
    3: {'name': 'Creativity & Presentation', 'max': 20, 'field': 'creativity'},
    4: {'name': 'Confidence & Presence', 'max': 15, 'field': 'confidence'},
    5: {'name': 'Overall Impact', 'max': 15, 'field': 'impact'}
}

CATEGORIES_COSTUME = {
    1: {'name': 'Creativity and Originality', 'max': 35, 'field': 'creativity'},
    2: {'name': 'Relevance to Department Theme', 'max': 30, 'field': 'relevance'},
    3: {'name': 'Stage Presence and Confidence', 'max': 20, 'field': 'presence'},
    4: {'name': 'Overall Impact and Presentation', 'max': 15, 'field': 'impact'}
}

CATEGORIES_SPORTS = {
    1: {'name': 'Confidence and Bearing', 'max': 30, 'field': 'confidence'},
    2: {'name': 'Style and Creativity of Sports Attire', 'max': 25, 'field': 'style'},
    3: {'name': 'Ramp Walk / Poise', 'max': 25, 'field': 'walk'},
    4: {'name': 'Audience Impact and Overall Presentation', 'max': 20, 'field': 'impact'}
}

CATEGORIES_QA = {
    1: {'name': 'Intelligence & Content', 'max': 40, 'field': 'intelligence'},
    2: {'name': 'Delivery & Wit', 'max': 40, 'field': 'delivery'},
    3: {'name': 'Overall Impact & Poise', 'max': 20, 'field': 'impact'}
}


# --- MODELS ---
class Candidate(db.Model):
    __tablename__ = 'hr_candidates'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(255), nullable=False)
    department = db.Column(db.String(100), nullable=False)


class Judge(db.Model):
    __tablename__ = 'hr_judges'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)


# Table 1: Original General Show Scores
class Score(db.Model):
    __tablename__ = 'hr_scores'
    id = db.Column(db.Integer, primary_key=True)
    candidate_id = db.Column(db.Integer, db.ForeignKey('hr_candidates.id'))
    judge_name = db.Column(db.String(100))
    dept_rep = db.Column(db.Float, default=0.0)
    comm_skills = db.Column(db.Float, default=0.0)
    creativity = db.Column(db.Float, default=0.0)
    confidence = db.Column(db.Float, default=0.0)
    impact = db.Column(db.Float, default=0.0)
    total_score = db.Column(db.Float, default=0.0)


# Table 2: Separate Creative Costume Scores
class ScoreCostume(db.Model):
    __tablename__ = 'hr_scores_costume'
    id = db.Column(db.Integer, primary_key=True)
    candidate_id = db.Column(db.Integer, db.ForeignKey('hr_candidates.id'))
    judge_name = db.Column(db.String(100))
    creativity = db.Column(db.Float, default=0.0)
    relevance = db.Column(db.Float, default=0.0)
    presence = db.Column(db.Float, default=0.0)
    impact = db.Column(db.Float, default=0.0)
    total_score = db.Column(db.Float, default=0.0)


# Table 3: Separate Sports Attire Welcome Ramp Scores
class ScoreSports(db.Model):
    __tablename__ = 'hr_scores_sports'
    id = db.Column(db.Integer, primary_key=True)
    candidate_id = db.Column(db.Integer, db.ForeignKey('hr_candidates.id'))
    judge_name = db.Column(db.String(100))
    confidence = db.Column(db.Float, default=0.0)
    style = db.Column(db.Float, default=0.0)
    walk = db.Column(db.Float, default=0.0)
    impact = db.Column(db.Float, default=0.0)
    total_score = db.Column(db.Float, default=0.0)


# Table 4: Separate Q&A Segment Scores
class ScoreQA(db.Model):
    __tablename__ = 'hr_scores_qa'
    id = db.Column(db.Integer, primary_key=True)
    candidate_id = db.Column(db.Integer, db.ForeignKey('hr_candidates.id'))
    judge_name = db.Column(db.String(100))
    intelligence = db.Column(db.Float, default=0.0)
    delivery = db.Column(db.Float, default=0.0)
    impact = db.Column(db.Float, default=0.0)
    total_score = db.Column(db.Float, default=0.0)


class EmployeeVote(db.Model):
    __tablename__ = 'hr_employee_votes'
    id = db.Column(db.Integer, primary_key=True)
    candidate_id = db.Column(db.Integer, db.ForeignKey('hr_candidates.id'))
    employee_id = db.Column(db.String(50), unique=True, nullable=False)


class Question(db.Model):
    __tablename__ = 'hr_questions'
    id = db.Column(db.Integer, primary_key=True)
    text = db.Column(db.Text, nullable=False)
    drawn = db.Column(db.Boolean, default=False)


# --- ROUTES ---

@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        code = request.form.get('access_code')
        if code == '6677':
            session['role'] = 'hr'
            return redirect(url_for('hr_results'))
        elif code == '8899':
            session['role'] = 'judge'
            session.pop('judge_name', None)
            return redirect(url_for('set_judge_name'))
        elif code == '1122':
            session['role'] = 'employee'
            return redirect(url_for('employee_poll'))
        else:
            flash("Invalid Access Code!", "danger")
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# --- JUDGES SELECTION GATEWAY ---

@app.route('/judge/set-name', methods=['GET', 'POST'])
def set_judge_name():
    if session.get('role') != 'judge': return redirect(url_for('login'))
    if request.method == 'POST':
        session['judge_name'] = request.form.get('judge_name')
        return redirect(url_for('select_round'))
    judges = Judge.query.order_by(Judge.name).all()
    return render_template('set_name.html', judges=judges)


@app.route('/judge/select-round')
def select_round():
    if session.get('role') != 'judge' or not session.get('judge_name'):
        return redirect(url_for('login'))
    return render_template('select_round.html')


# --- SEGMENT 1: GENERAL SHOW (ROUND 1) ---

@app.route('/judge/dashboard')
def judge_index():
    if session.get('role') != 'judge' or not session.get('judge_name'):
        return redirect(url_for('login'))

    judge_name = session.get('judge_name')
    total_candidates = Candidate.query.count() or 1

    progress = {}
    for i, cat in CATEGORIES.items():
        field = getattr(Score, cat['field'])
        count = Score.query.filter(Score.judge_name == judge_name, field > 0).count()
        progress[i] = {'name': cat['name'], 'count': count, 'total': total_candidates}

    return render_template('judge_index.html', progress=progress, round_title="Round 1: General Show",
                           score_endpoint='score_category')


@app.route('/judge/category/<int:cat_id>', methods=['GET', 'POST'])
def score_category(cat_id):
    if session.get('role') != 'judge' or not session.get('judge_name'):
        return redirect(url_for('login'))

    cat = CATEGORIES.get(cat_id)
    judge_name = session.get('judge_name')
    candidates = Candidate.query.all()

    if request.method == 'POST':
        for can in candidates:
            val = float(request.form.get(str(can.id), 0))
            score_row = Score.query.filter_by(candidate_id=can.id, judge_name=judge_name).first()
            if not score_row:
                score_row = Score(candidate_id=can.id, judge_name=judge_name)
                db.session.add(score_row)

            setattr(score_row, cat['field'], val)
            score_row.total_score = (score_row.dept_rep or 0) + (score_row.comm_skills or 0) + \
                                    (score_row.creativity or 0) + (score_row.confidence or 0) + \
                                    (score_row.impact or 0)

        db.session.commit()
        flash(f"Scores for {cat['name']} saved!", "success")
        return redirect(url_for('judge_index'))

    existing_scores = {s.candidate_id: getattr(s, cat['field'])
                       for s in Score.query.filter_by(judge_name=judge_name).all()}

    return render_template('score_category.html', cat=cat, cat_id=cat_id,
                           candidates=candidates, existing_scores=existing_scores, back_endpoint='judge_index')


# --- SEGMENT 2: CREATIVE COSTUME ---

@app.route('/judge/costume/dashboard')
def judge_index_costume():
    if session.get('role') != 'judge' or not session.get('judge_name'):
        return redirect(url_for('login'))

    judge_name = session.get('judge_name')
    total_candidates = Candidate.query.count() or 1

    progress = {}
    for i, cat in CATEGORIES_COSTUME.items():
        field = getattr(ScoreCostume, cat['field'])
        count = ScoreCostume.query.filter(ScoreCostume.judge_name == judge_name, field > 0).count()
        progress[i] = {'name': cat['name'], 'count': count, 'total': total_candidates}

    return render_template('judge_index.html', progress=progress, round_title="Best in Creative Costume",
                           score_endpoint='score_category_costume')


@app.route('/judge/costume/category/<int:cat_id>', methods=['GET', 'POST'])
def score_category_costume(cat_id):
    if session.get('role') != 'judge' or not session.get('judge_name'):
        return redirect(url_for('login'))

    cat = CATEGORIES_COSTUME.get(cat_id)
    judge_name = session.get('judge_name')
    candidates = Candidate.query.all()

    if request.method == 'POST':
        for can in candidates:
            val = float(request.form.get(str(can.id), 0))
            score_row = ScoreCostume.query.filter_by(candidate_id=can.id, judge_name=judge_name).first()
            if not score_row:
                score_row = ScoreCostume(candidate_id=can.id, judge_name=judge_name)
                db.session.add(score_row)

            setattr(score_row, cat['field'], val)
            score_row.total_score = (score_row.creativity or 0) + (score_row.relevance or 0) + \
                                    (score_row.presence or 0) + (score_row.impact or 0)

        db.session.commit()
        flash(f"Scores for {cat['name']} saved!", "success")
        return redirect(url_for('judge_index_costume'))

    existing_scores = {s.candidate_id: getattr(s, cat['field'])
                       for s in ScoreCostume.query.filter_by(judge_name=judge_name).all()}

    return render_template('score_category.html', cat=cat, cat_id=cat_id,
                           candidates=candidates, existing_scores=existing_scores, back_endpoint='judge_index_costume')


# --- SEGMENT 3: SPORTS ATTIRE ---

@app.route('/judge/sports/dashboard')
def judge_index_sports():
    if session.get('role') != 'judge' or not session.get('judge_name'):
        return redirect(url_for('login'))

    judge_name = session.get('judge_name')
    total_candidates = Candidate.query.count() or 1

    progress = {}
    for i, cat in CATEGORIES_SPORTS.items():
        field = getattr(ScoreSports, cat['field'])
        count = ScoreSports.query.filter(ScoreSports.judge_name == judge_name, field > 0).count()
        progress[i] = {'name': cat['name'], 'count': count, 'total': total_candidates}

    return render_template('judge_index.html', progress=progress, round_title="Welcome Ramp (Sports Attire)",
                           score_endpoint='score_category_sports')


@app.route('/judge/sports/category/<int:cat_id>', methods=['GET', 'POST'])
def score_category_sports(cat_id):
    if session.get('role') != 'judge' or not session.get('judge_name'):
        return redirect(url_for('login'))

    cat = CATEGORIES_SPORTS.get(cat_id)
    judge_name = session.get('judge_name')
    candidates = Candidate.query.all()

    if request.method == 'POST':
        for can in candidates:
            val = float(request.form.get(str(can.id), 0))
            score_row = ScoreSports.query.filter_by(candidate_id=can.id, judge_name=judge_name).first()
            if not score_row:
                score_row = ScoreSports(candidate_id=can.id, judge_name=judge_name)
                db.session.add(score_row)

            setattr(score_row, cat['field'], val)
            score_row.total_score = (score_row.confidence or 0) + (score_row.style or 0) + \
                                    (score_row.walk or 0) + (score_row.impact or 0)

        db.session.commit()
        flash(f"Scores for {cat['name']} saved!", "success")
        return redirect(url_for('judge_index_sports'))

    existing_scores = {s.candidate_id: getattr(s, cat['field'])
                       for s in ScoreSports.query.filter_by(judge_name=judge_name).all()}

    return render_template('score_category.html', cat=cat, cat_id=cat_id,
                           candidates=candidates, existing_scores=existing_scores, back_endpoint='judge_index_sports')


# --- SEGMENT 4: Q&A SEGMENT ---

@app.route('/judge/qa/dashboard')
def judge_index_qa():
    if session.get('role') != 'judge' or not session.get('judge_name'):
        return redirect(url_for('login'))

    judge_name = session.get('judge_name')
    total_candidates = Candidate.query.count() or 1

    progress = {}
    for i, cat in CATEGORIES_QA.items():
        field = getattr(ScoreQA, cat['field'])
        count = ScoreQA.query.filter(ScoreQA.judge_name == judge_name, field > 0).count()
        progress[i] = {'name': cat['name'], 'count': count, 'total': total_candidates}

    return render_template('judge_index.html', progress=progress, round_title="Q&A Segment",
                           score_endpoint='score_category_qa')


@app.route('/judge/qa/category/<int:cat_id>', methods=['GET', 'POST'])
def score_category_qa(cat_id):
    if session.get('role') != 'judge' or not session.get('judge_name'):
        return redirect(url_for('login'))

    cat = CATEGORIES_QA.get(cat_id)
    judge_name = session.get('judge_name')
    candidates = Candidate.query.all()

    if request.method == 'POST':
        for can in candidates:
            val = float(request.form.get(str(can.id), 0))
            score_row = ScoreQA.query.filter_by(candidate_id=can.id, judge_name=judge_name).first()
            if not score_row:
                score_row = ScoreQA(candidate_id=can.id, judge_name=judge_name)
                db.session.add(score_row)

            setattr(score_row, cat['field'], val)
            score_row.total_score = (score_row.intelligence or 0) + (score_row.delivery or 0) + \
                                    (score_row.impact or 0)

        db.session.commit()
        flash(f"Scores for {cat['name']} saved!", "success")
        return redirect(url_for('judge_index_qa'))

    existing_scores = {s.candidate_id: getattr(s, cat['field'])
                       for s in ScoreQA.query.filter_by(judge_name=judge_name).all()}

    return render_template('score_category.html', cat=cat, cat_id=cat_id,
                           candidates=candidates, existing_scores=existing_scores, back_endpoint='judge_index_qa')


# --- EMPLOYEE POLL ---

@app.route('/poll', methods=['GET', 'POST'])
def employee_poll():
    if session.get('role') != 'employee': return redirect(url_for('login'))
    if request.method == 'POST':
        emp_id = request.form.get('employee_id', '').strip().upper()
        can_id = request.form.get('candidate_id')
        if not emp_id:
            flash("ID required!", "danger")
            return redirect(url_for('employee_poll'))

        if EmployeeVote.query.filter_by(employee_id=emp_id).first():
            flash(f"ID {emp_id} already voted!", "danger")
            return redirect(url_for('employee_poll'))

        db.session.add(EmployeeVote(candidate_id=can_id, employee_id=emp_id))
        db.session.commit()
        return render_template('poll_success.html')

    candidates = Candidate.query.all()
    return render_template('employee_poll.html', candidates=candidates)


# --- HR DASHBOARDS ---

@app.route('/hr-results')
def hr_results():
    if session.get('role') != 'hr': return redirect(url_for('login'))
    # Overall dashboard is now the landing tabulation screen
    return redirect(url_for('hr_results_overall'))


@app.route('/hr-results/overall')
def hr_results_overall():
    if session.get('role') != 'hr': return redirect(url_for('login'))

    results = get_tabulated_results_overall()
    all_candidates = Candidate.query.order_by(Candidate.name).all()
    all_judges = Judge.query.order_by(Judge.name).all()

    return render_template('hr_results.html',
                           results=results,
                           all_candidates=all_candidates,
                           all_judges=all_judges,
                           total_emp_votes=0,
                           active_tab='overall',
                           reveal_endpoint='winner_reveal_overall',
                           export_endpoint='export_excel_overall',
                           wipe_endpoint='wipe_scores_overall')


@app.route('/hr-results/general')
def hr_results_general():
    if session.get('role') != 'hr': return redirect(url_for('login'))

    results, total_votes = get_tabulated_results()
    all_candidates = Candidate.query.order_by(Candidate.name).all()
    all_judges = Judge.query.order_by(Judge.name).all()

    return render_template('hr_results.html',
                           results=results,
                           all_candidates=all_candidates,
                           all_judges=all_judges,
                           total_emp_votes=total_votes,
                           active_tab='general',
                           reveal_endpoint='winner_reveal',
                           export_endpoint='export_excel',
                           wipe_endpoint='wipe_scores')


@app.route('/hr-results/costume')
def hr_results_costume():
    if session.get('role') != 'hr': return redirect(url_for('login'))

    results = get_tabulated_results_costume()
    all_candidates = Candidate.query.order_by(Candidate.name).all()
    all_judges = Judge.query.order_by(Judge.name).all()

    return render_template('hr_results.html',
                           results=results,
                           all_candidates=all_candidates,
                           all_judges=all_judges,
                           total_emp_votes=0,
                           active_tab='costume',
                           reveal_endpoint='winner_reveal_costume',
                           export_endpoint='export_excel_costume',
                           wipe_endpoint='wipe_scores_costume')


@app.route('/hr-results/sports')
def hr_results_sports():
    if session.get('role') != 'hr': return redirect(url_for('login'))

    results = get_tabulated_results_sports()
    all_candidates = Candidate.query.order_by(Candidate.name).all()
    all_judges = Judge.query.order_by(Judge.name).all()

    return render_template('hr_results.html',
                           results=results,
                           all_candidates=all_candidates,
                           all_judges=all_judges,
                           total_emp_votes=0,
                           active_tab='sports',
                           reveal_endpoint='winner_reveal_sports',
                           export_endpoint='export_excel_sports',
                           wipe_endpoint='wipe_scores_sports')


@app.route('/hr-results/qa')
def hr_results_qa():
    if session.get('role') != 'hr': return redirect(url_for('login'))

    results = get_tabulated_results_qa()
    all_candidates = Candidate.query.order_by(Candidate.name).all()
    all_judges = Judge.query.order_by(Judge.name).all()

    return render_template('hr_results.html',
                           results=results,
                           all_candidates=all_candidates,
                           all_judges=all_judges,
                           total_emp_votes=0,
                           active_tab='qa',
                           reveal_endpoint='winner_reveal_qa',
                           export_endpoint='export_excel_qa',
                           wipe_endpoint='wipe_scores_qa')


# --- MANAGEMENT ---

@app.route('/hr/add-candidate', methods=['GET', 'POST'])
def add_candidate():
    if session.get('role') != 'hr': return redirect(url_for('login'))
    if request.method == 'POST':
        db.session.add(Candidate(name=request.form['name'], department=request.form['dept']))
        db.session.commit()
        return redirect(url_for('hr_results'))
    candidates = Candidate.query.order_by(Candidate.name).all()
    return render_template('add_candidate.html', candidates=candidates)


@app.route('/hr/add-judge', methods=['GET', 'POST'])
def add_judge():
    if session.get('role') != 'hr': return redirect(url_for('login'))
    if request.method == 'POST':
        db.session.add(Judge(name=request.form['name']))
        db.session.commit()
        return redirect(url_for('hr_results'))
    judges = Judge.query.order_by(Judge.name).all()
    return render_template('add_judge.html', judges=judges)


@app.route('/hr/delete-candidate/<int:id>', methods=['POST'])
def delete_candidate(id):
    if session.get('role') != 'hr': return redirect(url_for('login'))
    Score.query.filter_by(candidate_id=id).delete()
    ScoreCostume.query.filter_by(candidate_id=id).delete()
    ScoreSports.query.filter_by(candidate_id=id).delete()
    ScoreQA.query.filter_by(candidate_id=id).delete()
    EmployeeVote.query.filter_by(candidate_id=id).delete()
    db.session.delete(Candidate.query.get_or_404(id))
    db.session.commit()
    return redirect(url_for('hr_results'))


@app.route('/hr/delete-judge/<int:id>', methods=['POST'])
def delete_judge(id):
    if session.get('role') != 'hr': return redirect(url_for('login'))
    judge = Judge.query.get_or_404(id)
    Score.query.filter_by(judge_name=judge.name).delete()
    ScoreCostume.query.filter_by(judge_name=judge.name).delete()
    ScoreSports.query.filter_by(judge_name=judge.name).delete()
    ScoreQA.query.filter_by(judge_name=judge.name).delete()
    db.session.delete(judge)
    db.session.commit()
    return redirect(url_for('hr_results'))


# --- WIPING UTILITIES ---

@app.route('/hr/wipe-scores', methods=['POST'])
def wipe_scores():
    if session.get('role') != 'hr': return redirect(url_for('login'))
    if request.form.get('wipe_password') == 'hr@55':
        db.session.query(Score).delete()
        db.session.query(EmployeeVote).delete()
        db.session.commit()
        flash("All Round 1 scores and votes wiped!", "success")
    return redirect(url_for('hr_results_general'))


@app.route('/hr/costume/wipe-scores', methods=['POST'])
def wipe_scores_costume():
    if session.get('role') != 'hr': return redirect(url_for('login'))
    if request.form.get('wipe_password') == 'hr@55':
        db.session.query(ScoreCostume).delete()
        db.session.commit()
        flash("All Creative Costume scores wiped!", "success")
    return redirect(url_for('hr_results_costume'))


@app.route('/hr/sports/wipe-scores', methods=['POST'])
def wipe_scores_sports():
    if session.get('role') != 'hr': return redirect(url_for('login'))
    if request.form.get('wipe_password') == 'hr@55':
        db.session.query(ScoreSports).delete()
        db.session.commit()
        flash("All Sports Attire scores wiped!", "success")
    return redirect(url_for('hr_results_sports'))


@app.route('/hr/qa/wipe-scores', methods=['POST'])
def wipe_scores_qa():
    if session.get('role') != 'hr': return redirect(url_for('login'))
    if request.form.get('wipe_password') == 'hr@55':
        db.session.query(ScoreQA).delete()
        db.session.commit()
        flash("All Q&A Segment scores wiped!", "success")
    return redirect(url_for('hr_results_qa'))


@app.route('/hr/overall/wipe-scores', methods=['POST'])
def wipe_scores_overall():
    if session.get('role') != 'hr': return redirect(url_for('login'))
    if request.form.get('wipe_password') == 'hr@55':
        db.session.query(Score).delete()
        db.session.query(ScoreCostume).delete()
        db.session.query(ScoreSports).delete()
        db.session.query(ScoreQA).delete()
        db.session.query(EmployeeVote).delete()
        db.session.commit()
        flash("System Reset: All scores across all rounds have been wiped!", "success")
    return redirect(url_for('hr_results_overall'))


# --- STAGE REVEALS ---

@app.route('/hr/reveal')
def winner_reveal():
    if session.get('role') != 'hr': return redirect(url_for('login'))
    results, _ = get_tabulated_results()
    return render_template('winner_reveal.html', winners=results)


@app.route('/hr/costume/reveal')
def winner_reveal_costume():
    if session.get('role') != 'hr': return redirect(url_for('login'))
    results = get_tabulated_results_costume()
    return render_template('winner_reveal.html', winners=results)


@app.route('/hr/sports/reveal')
def winner_reveal_sports():
    if session.get('role') != 'hr': return redirect(url_for('login'))
    results = get_tabulated_results_sports()
    return render_template('winner_reveal.html', winners=results)


@app.route('/hr/qa/reveal')
def winner_reveal_qa():
    if session.get('role') != 'hr': return redirect(url_for('login'))
    results = get_tabulated_results_qa()
    return render_template('winner_reveal.html', winners=results)


@app.route('/hr/overall/reveal')
def winner_reveal_overall():
    if session.get('role') != 'hr': return redirect(url_for('login'))
    results = get_tabulated_results_overall()
    return render_template('winner_reveal.html', winners=results)


# --- EXCEL EXPORTS ---

@app.route('/hr/export-excel')
def export_excel():
    if session.get('role') != 'hr': return redirect(url_for('login'))

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Final Tabulation R1"

    headers1 = [
        "Rank", "Candidate Name", "Department",
        "Avg Dept. Representation (Max 25)", "Avg Intro & Comm (Max 25)",
        "Avg Creativity (Max 20)", "Avg Confidence (Max 15)",
        "Avg Overall Impact (Max 15)", "Final Score (Avg Total)", "Judges Count"
    ]
    ws1.append(headers1)

    results, _ = get_tabulated_results()
    for rank, res in enumerate(results, start=1):
        scores = Score.query.filter_by(candidate_id=res['id']).all()
        judge_count = len(scores)

        if judge_count > 0:
            avg_dept = sum(s.dept_rep or 0.0 for s in scores) / judge_count
            avg_comm = sum(s.comm_skills or 0.0 for s in scores) / judge_count
            avg_creat = sum(s.creativity or 0.0 for s in scores) / judge_count
            avg_conf = sum(s.confidence or 0.0 for s in scores) / judge_count
            avg_imp = sum(s.impact or 0.0 for s in scores) / judge_count
        else:
            avg_dept = avg_comm = avg_creat = avg_conf = avg_imp = 0.0

        ws1.append([
            rank, res['name'], res['dept'],
            round(avg_dept, 2), round(avg_comm, 2), round(avg_creat, 2),
            round(avg_conf, 2), round(avg_imp, 2), res['final_score'], res['judge_count']
        ])

    ws2 = wb.create_sheet(title="Detailed Ballots R1")
    headers2 = [
        "Candidate Name", "Department", "Judge Name",
        "Dept. Representation", "Intro & Communication",
        "Creativity & Style", "Stage Poise", "Overall Impact", "Total Score"
    ]
    ws2.append(headers2)

    all_scores = db.session.query(
        Candidate.name, Candidate.department, Score.judge_name,
        Score.dept_rep, Score.comm_skills, Score.creativity,
        Score.confidence, Score.impact, Score.total_score
    ).join(Score, Candidate.id == Score.candidate_id).order_by(Candidate.name, Score.judge_name).all()

    for s in all_scores:
        ws2.append([s[0], s[1], s[2], s[3], s[4], s[5], s[6], s[7], s[8]])

    format_sheets(wb, [ws1, ws2])
    return generate_download(wb, "Binibining_Coolaire_2026_Round1_Tabulations.xlsx")


@app.route('/hr/costume/export-excel')
def export_excel_costume():
    if session.get('role') != 'hr': return redirect(url_for('login'))

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Costume Tabulation"

    headers1 = [
        "Rank", "Candidate Name", "Department",
        "Avg Creativity & Originality (Max 35)", "Avg Theme Relevance (Max 30)",
        "Avg Stage Presence (Max 20)", "Avg Overall Impact (Max 15)",
        "Final Score (Avg Total)", "Judges Count"
    ]
    ws1.append(headers1)

    results = get_tabulated_results_costume()
    for rank, res in enumerate(results, start=1):
        scores = ScoreCostume.query.filter_by(candidate_id=res['id']).all()
        judge_count = len(scores)

        if judge_count > 0:
            avg_cr = sum(s.creativity or 0.0 for s in scores) / judge_count
            avg_re = sum(s.relevance or 0.0 for s in scores) / judge_count
            avg_pr = sum(s.presence or 0.0 for s in scores) / judge_count
            avg_im = sum(s.impact or 0.0 for s in scores) / judge_count
        else:
            avg_cr = avg_re = avg_pr = avg_im = 0.0

        ws1.append([
            rank, res['name'], res['dept'],
            round(avg_cr, 2), round(avg_re, 2), round(avg_pr, 2), round(avg_im, 2),
            res['final_score'], res['judge_count']
        ])

    ws2 = wb.create_sheet(title="Costume Detailed Ballots")
    headers2 = [
        "Candidate Name", "Department", "Judge Name",
        "Creativity & Originality", "Theme Relevance",
        "Stage Presence", "Overall Impact", "Total Score"
    ]
    ws2.append(headers2)

    all_scores = db.session.query(
        Candidate.name, Candidate.department, ScoreCostume.judge_name,
        ScoreCostume.creativity, ScoreCostume.relevance, ScoreCostume.presence,
        ScoreCostume.impact, ScoreCostume.total_score
    ).join(ScoreCostume, Candidate.id == ScoreCostume.candidate_id).order_by(Candidate.name,
                                                                             ScoreCostume.judge_name).all()

    for s in all_scores:
        ws2.append([s[0], s[1], s[2], s[3], s[4], s[5], s[6], s[7]])

    format_sheets(wb, [ws1, ws2])
    return generate_download(wb, "Binibining_Coolaire_2026_Costume_Tabulations.xlsx")


@app.route('/hr/sports/export-excel')
def export_excel_sports():
    if session.get('role') != 'hr': return redirect(url_for('login'))

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sports Tabulation"

    headers1 = [
        "Rank", "Candidate Name", "Department",
        "Avg Confidence & Bearing (Max 30)", "Avg Style & Creativity (Max 25)",
        "Avg Ramp Walk (Max 25)", "Avg Audience Impact (Max 20)",
        "Final Score (Avg Total)", "Judges Count"
    ]
    ws1.append(headers1)

    results = get_tabulated_results_sports()
    for rank, res in enumerate(results, start=1):
        scores = ScoreSports.query.filter_by(candidate_id=res['id']).all()
        judge_count = len(scores)

        if judge_count > 0:
            avg_co = sum(s.confidence or 0.0 for s in scores) / judge_count
            avg_st = sum(s.style or 0.0 for s in scores) / judge_count
            avg_wa = sum(s.walk or 0.0 for s in scores) / judge_count
            avg_im = sum(s.impact or 0.0 for s in scores) / judge_count
        else:
            avg_co = avg_st = avg_wa = avg_im = 0.0

        ws1.append([
            rank, res['name'], res['dept'],
            round(avg_co, 2), round(avg_st, 2), round(avg_wa, 2), round(avg_im, 2),
            res['final_score'], res['judge_count']
        ])

    ws2 = wb.create_sheet(title="Sports Detailed Ballots")
    headers2 = [
        "Candidate Name", "Department", "Judge Name",
        "Confidence & Bearing", "Style & Creativity",
        "Ramp Walk / Poise", "Audience Impact", "Total Score"
    ]
    ws2.append(headers2)

    all_scores = db.session.query(
        Candidate.name, Candidate.department, ScoreSports.judge_name,
        ScoreSports.confidence, ScoreSports.style, ScoreSports.walk,
        ScoreSports.impact, ScoreSports.total_score
    ).join(ScoreSports, Candidate.id == ScoreSports.candidate_id).order_by(Candidate.name, ScoreSports.judge_name).all()

    for s in all_scores:
        ws2.append([s[0], s[1], s[2], s[3], s[4], s[5], s[6], s[7]])

    format_sheets(wb, [ws1, ws2])
    return generate_download(wb, "Binibining_Coolaire_2026_Sports_Tabulations.xlsx")


@app.route('/hr/qa/export-excel')
def export_excel_qa():
    if session.get('role') != 'hr': return redirect(url_for('login'))

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Q&A Tabulation"

    headers1 = [
        "Rank", "Candidate Name", "Department",
        "Avg Intelligence & Content (Max 40)", "Avg Delivery & Wit (Max 40)",
        "Avg Overall Impact & Poise (Max 20)",
        "Final Score (Avg Total)", "Judges Count"
    ]
    ws1.append(headers1)

    results = get_tabulated_results_qa()
    for rank, res in enumerate(results, start=1):
        scores = ScoreQA.query.filter_by(candidate_id=res['id']).all()
        judge_count = len(scores)

        if judge_count > 0:
            avg_in = sum(s.intelligence or 0.0 for s in scores) / judge_count
            avg_de = sum(s.delivery or 0.0 for s in scores) / judge_count
            avg_im = sum(s.impact or 0.0 for s in scores) / judge_count
        else:
            avg_in = avg_de = avg_im = 0.0

        ws1.append([
            rank, res['name'], res['dept'],
            round(avg_in, 2), round(avg_de, 2), round(avg_im, 2),
            res['final_score'], res['judge_count']
        ])

    ws2 = wb.create_sheet(title="Q&A Detailed Ballots")
    headers2 = [
        "Candidate Name", "Department", "Judge Name",
        "Intelligence & Content", "Delivery & Wit",
        "Overall Impact & Poise", "Total Score"
    ]
    ws2.append(headers2)

    all_scores = db.session.query(
        Candidate.name, Candidate.department, ScoreQA.judge_name,
        ScoreQA.intelligence, ScoreQA.delivery, ScoreQA.impact,
        ScoreQA.total_score
    ).join(ScoreQA, Candidate.id == ScoreQA.candidate_id).order_by(Candidate.name, ScoreQA.judge_name).all()

    for s in all_scores:
        ws2.append([s[0], s[1], s[2], s[3], s[4], s[5], s[6]])

    format_sheets(wb, [ws1, ws2])
    return generate_download(wb, "Binibining_Coolaire_2026_QA_Tabulations.xlsx")


@app.route('/hr/overall/export-excel')
def export_excel_overall():
    if session.get('role') != 'hr': return redirect(url_for('login'))

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Overall Final Tabulation"

    headers1 = [
        "Rank", "Candidate Name", "Department",
        "Round 1 Avg (General) %",
        "Round 2 Avg (Costume) %",
        "Round 2 Avg (Sports) %",
        "Round 2 Avg (Q&A) %",
        "Final Combined Score (Avg Total %)"
    ]
    ws1.append(headers1)

    results = get_tabulated_results_overall()
    for rank, res in enumerate(results, start=1):
        ws1.append([
            rank, res['name'], res['dept'],
            res['r1'], res['costume'], res['sports'], res['qa'],
            res['final_score']
        ])

    format_sheets(wb, [ws1])
    return generate_download(wb, "Binibining_Coolaire_2026_Overall_Final_Tabulations.xlsx")


def format_sheets(wb, sheets):
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )

    for ws in sheets:
        ws.views.sheetView[0].showGridLines = True
        ws.row_dimensions[1].height = 28

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


def generate_download(wb, filename):
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


# --- Q&A SEGMENT ENDPOINTS ---

@app.route('/hr/questions', methods=['GET', 'POST'])
def hr_questions():
    if session.get('role') != 'hr': return redirect(url_for('login'))

    if request.method == 'POST':
        question_text = request.form.get('question_text', '').strip()
        if question_text:
            db.session.add(Question(text=question_text))
            db.session.commit()
            flash("New question added to the bank!", "success")
        return redirect(url_for('hr_questions'))

    questions = Question.query.order_by(Question.id.desc()).all()
    return render_template('hr_questions.html', questions=questions)


@app.route('/hr/questions/edit/<int:id>', methods=['POST'])
def edit_question(id):
    if session.get('role') != 'hr': return redirect(url_for('login'))
    question = Question.query.get_or_404(id)
    question_text = request.form.get('question_text', '').strip()
    if question_text:
        question.text = question_text
        db.session.commit()
        flash("Question updated successfully!", "success")
    return redirect(url_for('hr_questions'))


@app.route('/hr/questions/delete/<int:id>', methods=['POST'])
def delete_question(id):
    if session.get('role') != 'hr': return redirect(url_for('login'))
    question = Question.query.get_or_404(id)
    db.session.delete(question)
    db.session.commit()
    flash("Question deleted.", "success")
    return redirect(url_for('hr_questions'))


@app.route('/hr/questions/reset-drawn', methods=['POST'])
def reset_drawn_questions():
    if session.get('role') != 'hr': return redirect(url_for('login'))
    db.session.query(Question).update({Question.drawn: False})
    db.session.commit()
    flash("All questions marked as undrawn!", "success")
    return redirect(url_for('hr_questions'))


@app.route('/stage/qa')
def stage_qa():
    return render_template('stage_qa.html')


@app.route('/stage/qa/draw-api', methods=['POST'])
def draw_question_api():
    question = Question.query.filter_by(drawn=False).order_by(func.random()).first()

    if not question:
        return {"status": "empty", "message": "All questions have been drawn! Reset the deck from the Q&A Admin Bank."}

    question.drawn = True
    db.session.commit()

    return {
        "status": "success",
        "id": question.id,
        "text": question.text
    }


# --- MASTER TABULATION LOGICS ---

def get_tabulated_results():
    actual_vote_count = db.session.query(func.count(EmployeeVote.id)).scalar() or 0

    raw_data = db.session.query(
        Candidate.id, Candidate.name, Candidate.department,
        func.avg(Score.total_score).label('judge_avg'),
        func.count(func.distinct(Score.judge_name)).label('judge_count')
    ).join(Score, isouter=True).group_by(Candidate.id).all()

    processed = []
    for r in raw_data:
        can_id = r[0]
        j_avg = float(r[3]) if r[3] is not None else 0.0
        poll_count = db.session.query(func.count(EmployeeVote.id)).filter_by(candidate_id=can_id).scalar() or 0

        processed.append({
            'id': can_id, 'name': r[1], 'dept': r[2],
            'judge_avg': round(j_avg, 2), 'judge_count': r[4],
            'poll_count': poll_count, 'final_score': round(j_avg, 2)
        })

    processed.sort(key=lambda x: x['judge_avg'], reverse=True)
    return processed, actual_vote_count


def get_tabulated_results_costume():
    raw_data = db.session.query(
        Candidate.id, Candidate.name, Candidate.department,
        func.avg(ScoreCostume.total_score).label('judge_avg'),
        func.count(func.distinct(ScoreCostume.judge_name)).label('judge_count')
    ).join(ScoreCostume, isouter=True).group_by(Candidate.id).all()

    processed = []
    for r in raw_data:
        can_id = r[0]
        j_avg = float(r[3]) if r[3] is not None else 0.0

        processed.append({
            'id': can_id, 'name': r[1], 'dept': r[2],
            'judge_avg': round(j_avg, 2), 'judge_count': r[4],
            'final_score': round(j_avg, 2)
        })

    processed.sort(key=lambda x: x['judge_avg'], reverse=True)
    return processed


def get_tabulated_results_sports():
    raw_data = db.session.query(
        Candidate.id, Candidate.name, Candidate.department,
        func.avg(ScoreSports.total_score).label('judge_avg'),
        func.count(func.distinct(ScoreSports.judge_name)).label('judge_count')
    ).join(ScoreSports, isouter=True).group_by(Candidate.id).all()

    processed = []
    for r in raw_data:
        can_id = r[0]
        j_avg = float(r[3]) if r[3] is not None else 0.0

        processed.append({
            'id': can_id, 'name': r[1], 'dept': r[2],
            'judge_avg': round(j_avg, 2), 'judge_count': r[4],
            'final_score': round(j_avg, 2)
        })

    processed.sort(key=lambda x: x['judge_avg'], reverse=True)
    return processed


def get_tabulated_results_qa():
    raw_data = db.session.query(
        Candidate.id, Candidate.name, Candidate.department,
        func.avg(ScoreQA.total_score).label('judge_avg'),
        func.count(func.distinct(ScoreQA.judge_name)).label('judge_count')
    ).join(ScoreQA, isouter=True).group_by(Candidate.id).all()

    processed = []
    for r in raw_data:
        can_id = r[0]
        j_avg = float(r[3]) if r[3] is not None else 0.0

        processed.append({
            'id': can_id, 'name': r[1], 'dept': r[2],
            'judge_avg': round(j_avg, 2), 'judge_count': r[4],
            'final_score': round(j_avg, 2)
        })

    processed.sort(key=lambda x: x['judge_avg'], reverse=True)
    return processed


def get_tabulated_results_overall():
    candidates = Candidate.query.order_by(Candidate.name).all()

    r1_list, _ = get_tabulated_results()
    r1_dict = {item['id']: item['judge_avg'] for item in r1_list}

    costume_dict = {item['id']: item['judge_avg'] for item in get_tabulated_results_costume()}
    sports_dict = {item['id']: item['judge_avg'] for item in get_tabulated_results_sports()}
    qa_dict = {item['id']: item['judge_avg'] for item in get_tabulated_results_qa()}

    processed = []
    for c in candidates:
        r1 = r1_dict.get(c.id, 0.0)
        costume = costume_dict.get(c.id, 0.0)
        sports = sports_dict.get(c.id, 0.0)
        qa = qa_dict.get(c.id, 0.0)

        final_score = (r1 + costume + sports + qa) / 4.0

        processed.append({
            'id': c.id,
            'name': c.name,
            'dept': c.department,
            'r1': r1,
            'costume': costume,
            'sports': sports,
            'qa': qa,
            'final_score': round(final_score, 2),
            'judge_count': "-"  # Overall tabulation spans multiple categories
        })

    processed.sort(key=lambda x: x['final_score'], reverse=True)
    return processed


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(host='0.0.0.0', port=8080, debug=True)